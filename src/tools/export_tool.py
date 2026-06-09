# ============================================================
# src/tools/export_tool.py
# Tool de exportação de conversa (Excel / PDF)
# ============================================================
import io
import re

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Alignment, Font, PatternFill


class ExportTool:
    """
    Exporta o histórico da conversa para Excel ou PDF.
    Retorna os bytes do arquivo gerado.
    """

    name        = "export_chat"
    description = "Exporta a conversa atual para Excel ou PDF."
    permission  = "allow"

    def __call__(self, format: str = "excel", messages: list | None = None) -> bytes:
        messages = messages or []
        if format == "excel":
            return self._to_excel(messages)
        return self._to_pdf(messages)

    # ── Excel ────────────────────────────────────────────────
    def _to_excel(self, messages: list) -> bytes:
        df = pd.DataFrame(messages)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Conversa", index=False)
            ws = writer.sheets["Conversa"]
            for cell in ws[1]:
                cell.fill = PatternFill(fill_type="solid", fgColor="D7E4BC")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in ws.columns:
                letter = col[0].column_letter
                ws.column_dimensions[letter].width = min(
                    max(max(len(str(c.value or "")) for c in col) + 2, 20), 60
                )
        buf.seek(0)
        return buf.getvalue()

    # ── PDF ──────────────────────────────────────────────────
    def _to_pdf(self, messages: list) -> bytes:
        def _clean(t: str) -> str:
            return re.sub(r"[^\x00-\x7F]+", "", t)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Oráculo Analista — Histórico", ln=True, align="C")
        pdf.ln(5)
        pdf.set_font("Arial", size=11)
        for m in messages:
            pdf.multi_cell(0, 8, f"{_clean(m.get('role','?').capitalize())}: {_clean(m.get('content',''))}")
        return pdf.output(dest="S").encode("latin-1", errors="ignore")
